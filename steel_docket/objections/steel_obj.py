import json
import glob,os
#stuff for pdf to txt file
import csv
import codecs
from openpyxl import load_workbook

#Set up header for csvData
csvData = [[] for i in range(1001)]
csvData[0] = ('Submitter Name', 'Country of Headquarters', 'Reg.Gov ID Number','HTSUS Code',
'Company requesting exclusion','Total Requested Annual Exclusion', 'Reason for Objection')

#Iterate through and open each file in folder
count = 1;
for fileName in glob.glob('*.xlsx'):
    filepath = os.path.abspath(fileName)
    print('filename: ', fileName)
    print('filepath: ', filepath)
    wb = load_workbook(filename = fileName, read_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    csvData[count].append(ws['F7'].value)  #submitterName
    csvData[count].append(ws['F12'].value) #Country of headquarters
    csvData[count].append(ws['F17'].value) #Reg.Gov ID Number
    csvData[count].append(ws['J17'].value) #HTSUS Code
    csvData[count].append(ws['O7'].value)  #Company Requesting exclusion
    csvData[count].append(ws['O17'].value) #Total Requested Annual Exclusion
    csvData[count].append(ws['D76'].value) #Reason for Objection
    count = count + 1
    if os.path.isfile(fileName):
        os.remove(fileName)

fileNameCSV = "Objections_From_BIS_2018_0006.csv"
with open(fileNameCSV, "w") as csvfile:
    CSVwrite = csv.writer(csvfile)
    CSVwrite.writerows(csvData)
    print("\nSubmits can be found in: ", fileNameCSV)
