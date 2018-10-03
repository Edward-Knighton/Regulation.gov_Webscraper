import json
import glob,os
#stuff for pdf to txt file
import csv
import codecs
from openpyxl import load_workbook

#Currently overwrites file if it is already in folder,
#Should probably change this

#Set up header for csvData
csvData = [[] for i in range(1001)]
csvData[0] = ('Submitter Name','Harmonized Tariff Schedule Code of the United States(HTSUS)', 'Country of Headquarters', 'Primary type of Steel Activity',
'Total Requested Annual Exclusion', 'Reason for Exclusion')



#Iterate through and open each file
count = 1;
for fileName in glob.glob('*.xlsx'):
    wb = load_workbook(filename = fileName)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    csvData[count].append(ws['F8'].value)  #submitterName
    csvData[count].append(ws['O6'].value)  #HTSUS
    csvData[count].append(ws['F13'].value) #Country of headquarters
    csvData[count].append(ws['G28'].value) #Primary type of steel activity
    csvData[count].append(ws['N28'].value) #Total Requested Annual Exclusion
    csvData[count].append(ws['D34'].value) #Reason for exclusion
    count = count + 1
    if os.path.isfile(fileName):
        os.remove(fileName)

fileNameCSV = "Exclusion_Requests_From_BIS_2018_0006.csv"
with open(fileNameCSV, "w") as csvfile:
    CSVwrite = csv.writer(csvfile)
    CSVwrite.writerows(csvData)
    print()
    print("Submits can be found in: ", fileNameCSV)
